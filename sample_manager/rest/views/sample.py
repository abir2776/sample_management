from io import BytesIO

import openpyxl
from django.core.files.base import ContentFile
from django.db import transaction
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl_image_loader import SheetImageLoader
from rest_framework import status
from rest_framework.exceptions import APIException
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.generics import (
    ListAPIView,
    ListCreateAPIView,
    RetrieveAPIView,
    RetrieveUpdateDestroyAPIView,
)
from rest_framework.permissions import OR, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from common.choices import Status
from sample_manager.choices import SampleStatus, StorageType
from sample_manager.models import GarmentSample, Image, SampleImage, Storage
from sample_manager.permissions import (
    IsAdministrator,
    IsSuperAdmin,
)
from sample_manager.rest.filters.sample_filter import GarmentSampleFilter
from sample_manager.rest.serializers.sample import (
    GarmentSampleHistorySerializer,
    SampleSerializer,
    SampleUploadSerializer,
)


class SampleListCreateView(ListCreateAPIView):
    serializer_class = SampleSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = GarmentSampleFilter
    search_fields = ["name", "style_no", "sku_no", "fabrication"]
    ordering_fields = ["name", "arrival_date", "color"]
    ordering = ["created_at"]

    def get_queryset(self):
        storage_uid = self.kwargs.get("storage_uid")
        storage = Storage.objects.filter(
            uid=storage_uid, type=StorageType.SPACE, status=Status.ACTIVE
        ).first()
        if not storage:
            raise APIException("Invalid storage uid provided")
        company = self.request.user.get_company()
        return GarmentSample.objects.filter(
            company=company,
            storage__uid=storage_uid,
            is_active=True,
            status=Status.ACTIVE,
        )

    def get_permissions(self):
        method = self.request.method

        if method == "GET":
            return [IsAuthenticated()]

        if method == "POST":
            return [IsAuthenticated()]

        return [IsAuthenticated()]


class SampleDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = SampleSerializer
    lookup_field = "uid"

    def get_permissions(self):
        method = self.request.method

        if method == "GET":
            return [IsAuthenticated()]

        if method in ["PUT", "PATCH"]:
            return [IsAuthenticated()]

        if method == "DELETE":
            return [OR(IsSuperAdmin(), IsAdministrator())]

        return [IsAuthenticated()]

    def get_queryset(self):
        storage_uid = self.kwargs.get("storage_uid")
        storage = Storage.objects.filter(
            uid=storage_uid, type=StorageType.SPACE, status=Status.ACTIVE
        ).first()
        if not storage:
            raise APIException("Invalid storage uid provided")
        company = self.request.user.get_company()
        return GarmentSample.objects.filter(
            company=company,
            storage__uid=storage_uid,
            is_active=True,
            status=Status.ACTIVE,
        )

    def delete(self, request, *args, **kwargs):
        sample = self.get_object()
        sample.status = Status.REMOVED
        sample.save()
        return Response(
            {"detail": "Sample deleted successfully"}, status=status.HTTP_204_NO_CONTENT
        )


class SampleListView(ListAPIView):
    serializer_class = SampleSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = GarmentSampleFilter
    search_fields = ["name", "style_no", "sku_no", "fabrication"]
    ordering_fields = ["name", "arrival_date", "color"]
    ordering = ["created_at"]

    def get_queryset(self):
        company = self.request.user.get_company()
        return GarmentSample.objects.filter(
            company=company,
            is_active=True,
            status=Status.ACTIVE,
        )

    def get_permissions(self):
        return [IsAuthenticated()]


class SampleSearchDetailView(RetrieveAPIView):
    serializer_class = SampleSerializer
    lookup_field = "uid"

    def get_permissions(self):
        return [IsAuthenticated()]

    def get_queryset(self):
        company = self.request.user.get_company()
        return GarmentSample.objects.filter(
            company=company,
            is_active=True,
            status=Status.ACTIVE,
        )


class GarmentSampleHistoryListView(ListAPIView):
    permission_classes = [OR(IsSuperAdmin(), IsAdministrator())]
    serializer_class = GarmentSampleHistorySerializer

    def get_queryset(self):
        uid = self.kwargs.get("uid")
        return GarmentSample.history.filter(id=uid).order_by("-history_date")


class SampleUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Upload Excel file to import garment samples

        Request Body:
        - file: Excel file (.xlsx or .xls)
        - storage_uid: Storage UID where samples will be stored

        Returns:
        - success: Boolean indicating overall success
        - message: Summary message
        - total_rows_processed: Number of rows processed
        - samples_created: Number of samples successfully created
        - samples_skipped: Number of samples skipped (duplicates/empty)
        - errors: Number of errors encountered
        - unique_colors: List of unique colors found
        - error_details: List of error details (if any)
        """
        serializer = SampleUploadSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        uploaded_file = serializer.validated_data["file"]
        storage_uid = serializer.validated_data["storage_uid"]

        user = request.user
        company = user.get_company()

        # Validate storage exists
        try:
            storage = Storage.objects.get(uid=storage_uid, type=StorageType.SPACE)
        except Storage.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": f"Storage with UID {storage_uid} not found",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        # Process the Excel file
        try:
            result = self.process_excel_file(uploaded_file, storage, user, company)

            response_data = {
                "success": result["errors"] == 0,
                "message": result["message"],
                "total_rows_processed": result["total_rows"],
                "samples_created": result["created"],
                "samples_skipped": result["skipped"],
                "errors": result["errors"],
                "unique_colors": sorted(list(result["unique_colors"])),
                "error_details": result["error_details"],
            }

            return Response(
                response_data,
                status=status.HTTP_201_CREATED
                if result["created"] > 0
                else status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"success": False, "message": f"Error processing file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def process_excel_file(self, file, storage, user, company):
        """Process uploaded Excel file and create samples"""
        created_count = 0
        skipped_count = 0
        error_count = 0
        unique_colors = set()
        error_details = []

        try:
            # Load workbook from uploaded file
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            image_loader = SheetImageLoader(sheet)
        except Exception as e:
            raise Exception(f"Error loading Excel file: {str(e)}")

        # Check if first row is header
        first_row_sample_id = (
            str(sheet["A1"].value).strip().upper() if sheet["A1"].value else ""
        )
        first_row_style = (
            str(sheet["B1"].value).strip().upper() if sheet["B1"].value else ""
        )

        header_keywords = [
            "SL",
            "NO",
            "SAMPLE",
            "STYLE",
            "PICTURE",
            "ITEM",
            "FABRIC",
            "GSM",
            "COLOUR",
            "COLOR",
            "SIZE",
        ]
        is_header_row = any(
            keyword in first_row_sample_id or keyword in first_row_style
            for keyword in header_keywords
        )

        start_row = 2 if is_header_row else 1
        total_rows = sheet.max_row - (start_row - 1)

        # Process each row
        for row_num in range(start_row, sheet.max_row + 1):
            try:
                with transaction.atomic():
                    # Extract data from cells
                    sample_id = sheet[f"A{row_num}"].value
                    style_no = sheet[f"B{row_num}"].value
                    name = sheet[f"D{row_num}"].value
                    fabrication = sheet[f"E{row_num}"].value
                    gsm_value = sheet[f"F{row_num}"].value
                    color = sheet[f"G{row_num}"].value
                    size_range = sheet[f"H{row_num}"].value

                    # Skip empty rows
                    if not sample_id and not style_no:
                        skipped_count += 1
                        continue

                    # Process color
                    color_str = str(color).strip() if color else ""
                    if color_str:
                        unique_colors.add(color_str)

                    # Check if sample already exists
                    if GarmentSample.objects.filter(
                        sample_id=sample_id, company=company
                    ).exists():
                        skipped_count += 1
                        error_details.append(
                            {
                                "row": row_num,
                                "sample_id": str(sample_id),
                                "error": "Sample already exists",
                            }
                        )
                        continue

                    # Create sample
                    sample = GarmentSample.objects.create(
                        storage=storage,
                        sample_id=str(sample_id) if sample_id else "",
                        created_by=user,
                        company=company,
                        style_no=str(style_no) if style_no else "",
                        name=str(name) if name else "",
                        fabrication=str(fabrication) if fabrication else "",
                        color=color_str,
                        size_range=str(size_range) if size_range else "",
                        status=SampleStatus.ACTIVE,
                        is_active=True,
                    )

                    # Handle image from Excel
                    picture_cell = f"C{row_num}"
                    if image_loader.image_in(picture_cell):
                        try:
                            image_data = image_loader.get(picture_cell)
                            image_filename = f"{sample_id}_{style_no}.png"

                            # Convert PIL Image to file
                            img_byte_arr = BytesIO()
                            image_data.save(img_byte_arr, format="PNG")
                            img_byte_arr.seek(0)

                            # Create Image object
                            image_obj = Image.objects.create(
                                company=company,
                                file_name=image_filename,
                                created_by=user,
                            )
                            image_obj.file.save(
                                image_filename,
                                ContentFile(img_byte_arr.getvalue()),
                                save=True,
                            )

                            # Link image to sample
                            SampleImage.objects.create(
                                company=company, sample=sample, image=image_obj
                            )

                        except Exception as img_error:
                            error_details.append(
                                {
                                    "row": row_num,
                                    "sample_id": str(sample_id),
                                    "error": f"Image processing failed: {str(img_error)}",
                                }
                            )

                    created_count += 1

            except Exception as e:
                error_count += 1
                error_details.append(
                    {
                        "row": row_num,
                        "sample_id": str(sample_id) if sample_id else "Unknown",
                        "error": str(e),
                    }
                )

        # Prepare result message
        if created_count > 0:
            message = f"Successfully imported {created_count} samples"
        else:
            message = "No samples were imported"

        if skipped_count > 0:
            message += f", skipped {skipped_count} duplicates/empty rows"

        if error_count > 0:
            message += f", encountered {error_count} errors"

        return {
            "created": created_count,
            "skipped": skipped_count,
            "errors": error_count,
            "unique_colors": unique_colors,
            "total_rows": total_rows,
            "message": message,
            "error_details": error_details,
        }
