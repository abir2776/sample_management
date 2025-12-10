"""
Django management command to import garment samples from Excel file(s)
Usage:
  Single file: python manage.py import_samples --file path/to/excel.xlsx --storage-uid <storage_uid> --user-id <user_id>
  Multiple files: python manage.py import_samples --files path/to/file1.xlsx path/to/file2.xlsx --storage-uid <storage_uid> --user-id <user_id>
  Directory: python manage.py import_samples --directory path/to/folder --storage-uid <storage_uid> --user-id <user_id>
"""

import os

import openpyxl
from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl_image_loader import SheetImageLoader

from core.models import User
from sample_manager.choices import SampleStatus, StorageType
from sample_manager.models import GarmentSample, Image, SampleImage, Storage


class Command(BaseCommand):
    help = "Import garment samples from Excel file(s) with images"

    def add_arguments(self, parser):
        parser.add_argument("--file", type=str, help="Path to single Excel file")
        parser.add_argument(
            "--files", nargs="+", type=str, help="Paths to multiple Excel files"
        )
        parser.add_argument(
            "--directory", type=str, help="Directory containing Excel files"
        )
        parser.add_argument(
            "--storage-uid", type=str, required=True, help="Storage UID"
        )
        parser.add_argument(
            "--user-id", type=int, required=True, help="User ID who is importing"
        )

    def handle(self, *args, **options):
        storage_uid = options["storage_uid"]
        user_id = options["user_id"]

        # Get user and company
        try:
            user = User.objects.get(id=user_id)
            company = user.get_company()
        except User.DoesNotExist:
            self.stdout.write(self.style.ERROR(f"User with ID {user_id} not found"))
            return

        # Get storage
        try:
            storage = Storage.objects.get(uid=storage_uid, type=StorageType.SPACE)
        except Storage.DoesNotExist:
            self.stdout.write(
                self.style.ERROR(f"Storage with UID {storage_uid} not found")
            )
            return

        # Collect file paths
        file_paths = []

        if options["file"]:
            file_paths.append(options["file"])
        elif options["files"]:
            file_paths.extend(options["files"])
        elif options["directory"]:
            directory = options["directory"]
            if not os.path.isdir(directory):
                self.stdout.write(self.style.ERROR(f"Directory not found: {directory}"))
                return
            # Get all Excel files from directory
            for filename in os.listdir(directory):
                if filename.endswith((".xlsx", ".xls")):
                    file_paths.append(os.path.join(directory, filename))
        else:
            self.stdout.write(
                self.style.ERROR("Please provide --file, --files, or --directory")
            )
            return

        if not file_paths:
            self.stdout.write(self.style.ERROR("No Excel files found to process"))
            return

        # Process each file
        total_created = 0
        total_skipped = 0
        total_errors = 0
        all_unique_colors = set()  # Track unique colors across all files

        for file_path in file_paths:
            self.stdout.write(self.style.SUCCESS(f"\n{'=' * 60}"))
            self.stdout.write(
                self.style.SUCCESS(f"Processing file: {os.path.basename(file_path)}")
            )
            self.stdout.write(self.style.SUCCESS(f"{'=' * 60}"))

            created, skipped, errors, unique_colors = self.process_file(
                file_path, storage, user, company
            )

            total_created += created
            total_skipped += skipped
            total_errors += errors
            all_unique_colors.update(unique_colors)

        # Overall Summary
        self.stdout.write(self.style.SUCCESS("\n" + "=" * 60))
        self.stdout.write(self.style.SUCCESS("OVERALL IMPORT SUMMARY:"))
        self.stdout.write(
            self.style.SUCCESS(f"Total files processed: {len(file_paths)}")
        )
        self.stdout.write(self.style.SUCCESS(f"Total samples created: {total_created}"))
        self.stdout.write(self.style.WARNING(f"Total skipped: {total_skipped}"))
        self.stdout.write(self.style.ERROR(f"Total errors: {total_errors}"))
        self.stdout.write(self.style.SUCCESS("=" * 60))

        # Print all unique colors
        if all_unique_colors:
            self.stdout.write(self.style.SUCCESS("\n" + "=" * 60))
            self.stdout.write(self.style.SUCCESS("ALL UNIQUE COLORS FOUND:"))
            self.stdout.write(self.style.SUCCESS("=" * 60))
            for color in sorted(all_unique_colors):
                if color:  # Skip empty colors
                    self.stdout.write(self.style.SUCCESS(f"  • {color}"))
            self.stdout.write(
                self.style.SUCCESS(
                    f"\nTotal unique colors: {len([c for c in all_unique_colors if c])}"
                )
            )
            self.stdout.write(self.style.SUCCESS("=" * 60))

    def process_file(self, file_path, storage, user, company):
        """Process a single Excel file"""
        created_count = 0
        error_count = 0
        skipped_count = 0
        unique_colors = set()  # Track unique colors in this file

        # Validate file exists
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return 0, 0, 1, unique_colors

        # Load workbook
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            image_loader = SheetImageLoader(sheet)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error loading Excel file: {str(e)}"))
            return 0, 0, 1, unique_colors

        # Define column mapping
        column_mapping = {
            "A": "sample_id",  # SL NO
            "B": "style_no",  # STYLE
            "C": "picture",  # PICTURE (image)
            "D": "name",  # ITEM (changed to name)
            "E": "fabrication",  # FABRIC
            "F": "gsm",  # GSM
            "G": "color",  # COLOUR
            "H": "size_range",  # SIZE RANGE
        }

        # Check if first row contains header by looking for common header keywords
        first_row_sample_id = (
            str(sheet["A1"].value).strip().upper() if sheet["A1"].value else ""
        )
        first_row_style = (
            str(sheet["B1"].value).strip().upper() if sheet["B1"].value else ""
        )

        # Common header keywords to detect
        header_keywords = [
            "SL",
            "NO",
            "SAMPLE",
            "STYLE",
            "PICTURE",
            "ITEM",
            "FABRIC",
            "GSM",
            "COLOUR",
            "COLOR",
            "SIZE",
        ]
        is_header_row = any(
            keyword in first_row_sample_id or keyword in first_row_style
            for keyword in header_keywords
        )

        # Start from row 2 if header exists, otherwise row 1
        start_row = 2 if is_header_row else 1

        if is_header_row:
            self.stdout.write(
                self.style.SUCCESS(
                    f"Header row detected, starting from row {start_row}"
                )
            )

        # Process rows
        for row_num in range(start_row, sheet.max_row + 1):
            try:
                with transaction.atomic():
                    # Extract data from cells
                    sample_id = sheet[f"A{row_num}"].value
                    style_no = sheet[f"B{row_num}"].value
                    name = sheet[f"D{row_num}"].value  # ITEM = name
                    fabrication = sheet[f"E{row_num}"].value
                    gsm_value = sheet[f"F{row_num}"].value
                    color = sheet[f"G{row_num}"].value
                    size_range = sheet[f"H{row_num}"].value

                    # Skip empty rows
                    if not sample_id and not style_no:
                        skipped_count += 1
                        continue

                    # Convert color to string and strip whitespace
                    color_str = str(color).strip() if color else ""

                    # Add color to unique colors set (including empty ones for tracking)
                    if color_str:
                        unique_colors.add(color_str)

                    # Check if sample already exists
                    if GarmentSample.objects.filter(
                        sample_id=sample_id, company=company
                    ).exists():
                        self.stdout.write(
                            self.style.WARNING(
                                f"Row {row_num}: Sample {sample_id} already exists, skipping"
                            )
                        )
                        skipped_count += 1
                        continue

                    # Create sample
                    sample = GarmentSample.objects.create(
                        storage=storage,
                        sample_id=str(sample_id) if sample_id else "",
                        created_by=user,
                        company=company,
                        style_no=str(style_no) if style_no else "",
                        name=str(name) if name else "",  # ITEM = name
                        fabrication=str(fabrication) if fabrication else "",
                        color=color_str,
                        size_range=str(size_range) if size_range else "",
                        status=SampleStatus.ACTIVE,
                        is_active=True,
                    )

                    # Handle image from Excel
                    picture_cell = f"C{row_num}"
                    if image_loader.image_in(picture_cell):
                        try:
                            # Get image from cell
                            image_data = image_loader.get(picture_cell)

                            # Save image to Image model
                            image_filename = f"{sample_id}_{style_no}.png"

                            # Convert PIL Image to file
                            from io import BytesIO

                            img_byte_arr = BytesIO()
                            image_data.save(img_byte_arr, format="PNG")
                            img_byte_arr.seek(0)

                            # Create Image object
                            image_obj = Image.objects.create(
                                company=company,
                                file_name=image_filename,
                                created_by=user,
                            )
                            image_obj.file.save(
                                image_filename,
                                ContentFile(img_byte_arr.getvalue()),
                                save=True,
                            )

                            # Link image to sample
                            SampleImage.objects.create(
                                company=company, sample=sample, image=image_obj
                            )

                        except Exception as img_error:
                            self.stdout.write(
                                self.style.WARNING(
                                    f"Row {row_num}: Could not process image - {str(img_error)}"
                                )
                            )

                    created_count += 1
                    self.stdout.write(
                        self.style.SUCCESS(
                            f"Row {row_num}: Created sample {sample_id} - {name}"
                        )
                    )

            except Exception as e:
                error_count += 1
                self.stdout.write(
                    self.style.ERROR(f"Row {row_num}: Error creating sample - {str(e)}")
                )

        # File Summary
        self.stdout.write(
            self.style.SUCCESS(f"\nFile Summary for {os.path.basename(file_path)}:")
        )
        self.stdout.write(
            self.style.SUCCESS(
                f"Total rows processed: {sheet.max_row - (start_row - 1)}"
            )
        )
        self.stdout.write(self.style.SUCCESS(f"Successfully created: {created_count}"))
        self.stdout.write(self.style.WARNING(f"Skipped: {skipped_count}"))
        self.stdout.write(self.style.ERROR(f"Errors: {error_count}"))

        # Print unique colors for this file
        if unique_colors:
            self.stdout.write(
                self.style.SUCCESS(
                    f"\nUnique colors in this file: {len(unique_colors)}"
                )
            )
            for color in sorted(unique_colors):
                self.stdout.write(self.style.SUCCESS(f"  • {color}"))

        return created_count, skipped_count, error_count, unique_colors
